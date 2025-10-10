from flask import Blueprint, request, jsonify, send_file
from services.salary_service import SalaryService
from models.employee import Employee
from models.wage_master import WageMaster
from models.attendance import Attendance
from models.deduction import Deduction
from models import db
from datetime import datetime, date
import pandas as pd
import io
import calendar

forms_bp = Blueprint("forms", __name__)

@forms_bp.route("/form-b", methods=["GET", "OPTIONS"])
def get_form_b_data():
    """
    OPTIMIZED: Get Form B (Wages Register) data using BULK queries instead of N+1
    Reduces from N*8 queries to ~10 total queries for any number of employees
    """
    # Handle preflight OPTIONS request
    if request.method == 'OPTIONS':
        return '', 200

    import time
    start_time = time.time()

    try:
        # Get query parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        site = request.args.get('site')

        if not year or not month:
            return jsonify({
                "success": False,
                "message": "Year and month are required"
            }), 400

        # ============================================
        # STEP 1: Get employees for the specified site (1 query)
        # ============================================
        query = Employee.query
        if site:
            # URL decode the site parameter
            from urllib.parse import unquote
            decoded_site = unquote(site)
            print(f"[DEBUG] Filtering by site: '{site}' -> '{decoded_site}'")

            # Join with WageMaster to filter by site
            query = query.join(WageMaster, Employee.salary_code == WageMaster.salary_code)\
                         .filter(WageMaster.site_name == decoded_site)

        # Apply sorting and execute query (for both site-filtered and all employees)
        employees = query.order_by(Employee.employee_id.asc()).all()

        if not employees:
            return jsonify({
                "success": False,
                "message": "No employees found for the specified criteria"
            }), 404

        employee_ids = [emp.employee_id for emp in employees]

        # ============================================
        # STEP 2: BULK calculate salaries for filtered employees (3 queries)
        # ============================================
        bulk_salary_start = time.time()
        # Pass the already filtered employee IDs to avoid double filtering
        bulk_salary_result = SalaryService.generate_monthly_salary_data_for_employees(employee_ids, year, month)
        bulk_salary_time = time.time() - bulk_salary_start

        if not bulk_salary_result['success']:
            return jsonify({
                'success': False,
                'message': f'Bulk salary calculation failed: {bulk_salary_result.get("message", "Unknown error")}'
            }), 500

        # Convert list to dictionary keyed by employee ID
        salary_data_list = bulk_salary_result['data']
        salary_data_dict = {item['Employee ID']: item for item in salary_data_list}

        # ============================================
        # STEP 3: BULK fetch attendance data for ALL employees (1 query)
        # ============================================
        bulk_attendance_start = time.time()
        from services.attendance_service import AttendanceService
        bulk_attendance_result = AttendanceService.get_bulk_monthly_attendance_summary(employee_ids, year, month)
        bulk_attendance_time = time.time() - bulk_attendance_start

        if not bulk_attendance_result['success']:
            return jsonify({
                'success': False,
                'message': f'Bulk attendance calculation failed: {bulk_attendance_result.get("message", "Unknown error")}'
            }), 500

        attendance_dict = bulk_attendance_result['data']

        # ============================================
        # STEP 4: BULK fetch deductions for ALL employees (1 query)
        # ============================================
        bulk_deductions_start = time.time()
        bulk_deductions_result = SalaryService.get_bulk_monthly_deductions(employee_ids, year, month)
        bulk_deductions_time = time.time() - bulk_deductions_start

        if not bulk_deductions_result['success']:
            return jsonify({
                'success': False,
                'message': f'Bulk deductions calculation failed: {bulk_deductions_result.get("message", "Unknown error")}'
            }), 500

        deductions_dict = bulk_deductions_result['data']

        # ============================================
        # STEP 5: BULK fetch wage master data for ALL employees (1 query)
        # ============================================
        wage_master_start = time.time()
        wage_master_query = WageMaster.query.filter(
            WageMaster.salary_code.in_([emp.salary_code for emp in employees if emp.salary_code])
        ).all()

        wage_master_dict = {wm.salary_code: wm for wm in wage_master_query}
        wage_master_time = time.time() - wage_master_start

        # ============================================
        # STEP 6: Build Form B rows from pre-calculated data (IN-MEMORY)
        # ============================================
        form_b_data = []

        for idx, employee in enumerate(employees, 1):
            # Get pre-calculated data from bulk queries
            salary_data = salary_data_dict.get(employee.employee_id, {})
            attendance_data = attendance_dict.get(employee.employee_id, {})
            deduction_data = deductions_dict.get(employee.employee_id, {})

            # Get wage master details
            wage_master = wage_master_dict.get(employee.salary_code)

            # Extract values from bulk data
            present_days = attendance_data.get('present_days', 0)
            overtime_hours = attendance_data.get('total_overtime_hours', 0)
            overtime_amount = salary_data.get('Overtime Allowance', 0)
            monthly_deduction_total = deduction_data.get('total_deduction', 0)

            # Map salary data to Form B structure
            form_b_row = {
                "slNo": idx,
                "employeeCode": employee.employee_id,
                "employeeName": f"{employee.first_name} {employee.last_name}",
                "designation": employee.designation or "N/A",
                "rateOfWage": {
                    "bs": wage_master.base_wage if wage_master else salary_data.get('Daily Wage', 0),
                    "da": 0  # DA component if available in adjustments
                },
                "daysWorked": present_days,
                "overtime": overtime_hours,
                "totalDays": present_days + (overtime_hours / 8),  # Convert OT hours to days
                "grossEarnings": {
                    "bs": salary_data.get('Basic', 0),
                    "da": salary_data.get('DA', 0),
                    "hra": salary_data.get('HRA', 0),
                    "cov": 0,  # Conveyance allowance
                    "ota": overtime_amount,
                    "ae": salary_data.get('Others', 0)  # Additional earnings
                },
                "totalEarnings": salary_data.get('Total Earnings', 0),
                "deductions": {
                    "pf": salary_data.get('PF', 0),
                    "esi": salary_data.get('ESIC', 0),
                    "cit": salary_data.get('Income Tax', 0),
                    "ptax": 0,  # Professional tax
                    "adv": salary_data.get('Others Recoveries', 0),
                    "otherRecoveries": monthly_deduction_total,  # New field for all deductions
                    "total": salary_data.get('Total Deductions', 0)
                },
                "netPayable": salary_data.get('Net Salary', 0),
                "siteName": wage_master.site_name if wage_master else "N/A"
            }

            form_b_data.append(form_b_row)

        # Calculate totals
        totals = {
            "totalEmployees": len(form_b_data),
            "totalDaysWorked": sum(row["daysWorked"] for row in form_b_data),
            "totalOvertime": sum(row["overtime"] for row in form_b_data),
            "totalEarnings": sum(row["totalEarnings"] for row in form_b_data),
            "totalDeductions": sum(row["deductions"]["total"] for row in form_b_data),
            "totalOtherRecoveries": sum(row["deductions"]["otherRecoveries"] or 0 for row in form_b_data),
            "totalNetPayable": sum(row["netPayable"] for row in form_b_data)
        }

        # Calculate performance metrics
        total_time = time.time() - start_time

        # Log performance metrics
        print(f"[PERFORMANCE] Form B generated in {total_time:.2f}s for {len(employees)} employees")
        print(f"[PERFORMANCE] Bulk salary: {bulk_salary_time:.3f}s, Attendance: {bulk_attendance_time:.3f}s, Deductions: {bulk_deductions_time:.3f}s, WageMaster: {wage_master_time:.3f}s")
        print(f"[PERFORMANCE] Total queries: ~6 (was {len(employees)} * 8 = {len(employees) * 8})")

        response = jsonify({
            "success": True,
            "data": form_b_data,
            "totals": totals,
            "filters": {
                "year": year,
                "month": month,
                "site": site,
                "monthName": calendar.month_name[month]
            },
            "performance": {
                "total_time_seconds": round(total_time, 3),
                "employees_processed": len(employees),
                "bulk_salary_time": round(bulk_salary_time, 3),
                "bulk_attendance_time": round(bulk_attendance_time, 3),
                "bulk_deductions_time": round(bulk_deductions_time, 3),
                "wage_master_time": round(wage_master_time, 3),
                "total_queries": 6,  # Approximate
                "old_queries_estimate": len(employees) * 8
            }
        })

        # Add performance headers
        response.headers['X-Performance-Total'] = f'{total_time:.3f}s'
        response.headers['X-Performance-Employees'] = str(len(employees))
        response.headers['X-Performance-Queries'] = '6'

        return response, 200

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error generating Form B data: {str(e)}"
        }), 500


@forms_bp.route("/form-b/download", methods=["GET", "OPTIONS"])
def download_form_b_excel():
    """
    Download Form B data as Excel file
    """
    # Handle preflight OPTIONS request
    if request.method == 'OPTIONS':
        return '', 200

    try:
        # Get query parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        site = request.args.get('site', 'All')

        if not year or not month:
            return jsonify({
                "success": False,
                "message": "Year and month are required"
            }), 400

        # Get Form B data using the same logic as the data endpoint
        form_b_response = get_form_b_data()
        
        if form_b_response[1] != 200:  # Check status code
            return form_b_response
            
        form_b_result = form_b_response[0].get_json()
        form_b_data = form_b_result['data']
        totals = form_b_result['totals']
        
        # Create Excel file
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Prepare data for Excel
            excel_data = []
            
            for row in form_b_data:
                excel_row = {
                    'Sl.No': row['slNo'],
                    'Employee Code': row['employeeCode'],
                    'Employee Name': row['employeeName'],
                    'Designation': row['designation'],
                    'Rate of Wage (BS)': row['rateOfWage']['bs'],
                    'Rate of Wage (DA)': row['rateOfWage']['da'],
                    'Days Worked': row['daysWorked'],
                    'Overtime': row['overtime'],
                    'Total Days': row['totalDays'],
                    'BS': row['grossEarnings']['bs'],
                    'DA': row['grossEarnings']['da'],
                    'HRA': row['grossEarnings']['hra'],
                    'COV': row['grossEarnings']['cov'],
                    'OTA': row['grossEarnings']['ota'],
                    'AE': row['grossEarnings']['ae'],
                    'Total Earnings': row['totalEarnings'],
                    'PF': row['deductions']['pf'],
                    'ESI': row['deductions']['esi'],
                    'CIT': row['deductions']['cit'],
                    'PTAX': row['deductions']['ptax'],
                    'ADV': row['deductions']['adv'],
                    'Other Recoveries': row['deductions']['otherRecoveries'],
                    'Total Deductions': row['deductions']['total'],
                    'Net Payable': row['netPayable']
                }
                excel_data.append(excel_row)
            
            # Add totals row
            totals_row = {
                'Sl.No': '',
                'Employee Code': '',
                'Employee Name': 'TOTAL',
                'Designation': '',
                'Rate of Wage (BS)': '',
                'Rate of Wage (DA)': '',
                'Days Worked': totals['totalDaysWorked'],
                'Overtime': totals['totalOvertime'],
                'Total Days': '',
                'BS': '',
                'DA': '',
                'HRA': '',
                'COV': '',
                'OTA': '',
                'AE': '',
                'Total Earnings': totals['totalEarnings'],
                'PF': '',
                'ESI': '',
                'CIT': '',
                'PTAX': '',
                'ADV': '',
                'Other Recoveries': totals['totalOtherRecoveries'],
                'Total Deductions': totals['totalDeductions'],
                'Net Payable': totals['totalNetPayable']
            }
            excel_data.append(totals_row)
            
            # Create DataFrame and write to Excel
            df = pd.DataFrame(excel_data)
            df.to_excel(writer, sheet_name='Form B - Wages Register', index=False)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Form B - Wages Register']
            
            # Add header information as requested by user
            worksheet.insert_rows(1, 5)  # Insert 5 rows for headers
            worksheet['A1'] = "Form B"
            worksheet['A2'] = "Format For Wage Register"
            worksheet['A3'] = f"Rate of minimum Wages {datetime.now().strftime('%d/%m/%Y')}"
            worksheet['A4'] = "SSPL"
            worksheet['A5'] = f"Month: {calendar.month_name[month]} {year} | Site: {site}"

            # Style the headers
            from openpyxl.styles import Font, Alignment

            # Make headers bold and centered
            for row in range(1, 6):
                cell = worksheet[f'A{row}']
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center')

            # Merge cells for headers to span across columns
            worksheet.merge_cells('A1:X1')  # Form B
            worksheet.merge_cells('A2:X2')  # Format For Wage Register
            worksheet.merge_cells('A3:X3')  # Rate of minimum wages with date
            worksheet.merge_cells('A4:X4')  # SSPL
            worksheet.merge_cells('A5:X5')  # Month and Site info
        
        output.seek(0)
        
        # Generate filename
        month_name = calendar.month_name[month]
        filename = f"FormB_{site}_{month_name}_{year}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error generating Excel file: {str(e)}"
        }), 500


@forms_bp.route("/form-d", methods=["GET", "OPTIONS"])
def get_form_d_data():
    """
    Get Form D (ESIC) data with wage ceilings and IP contribution calculations
    """
    # Handle preflight OPTIONS request
    if request.method == 'OPTIONS':
        return '', 200

    import time
    start_time = time.time()

    try:
        # Get query parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        site = request.args.get('site')

        if not year or not month:
            return jsonify({
                "success": False,
                "message": "Year and month are required"
            }), 400

        # ESIC wage ceiling (₹21,000)
        ESIC_WAGE_CEILING = 21000

        # ============================================
        # STEP 1: Get employees for the specified site (1 query)
        # ============================================
        query = Employee.query
        if site:
            # URL decode the site parameter
            from urllib.parse import unquote
            decoded_site = unquote(site)
            print(f"[DEBUG] Filtering by site: '{site}' -> '{decoded_site}'")

            # Join with WageMaster to filter by site
            query = query.join(WageMaster, Employee.salary_code == WageMaster.salary_code)\
                         .filter(WageMaster.site_name == decoded_site)

        # Apply sorting and execute query
        employees = query.order_by(Employee.employee_id.asc()).all()

        if not employees:
            return jsonify({
                "success": False,
                "message": "No employees found for the specified criteria"
            }), 404

        employee_ids = [emp.employee_id for emp in employees]

        # ============================================
        # STEP 2: BULK calculate salaries for filtered employees (3 queries)
        # ============================================
        bulk_salary_start = time.time()
        bulk_salary_result = SalaryService.generate_monthly_salary_data_for_employees(employee_ids, year, month)
        bulk_salary_time = time.time() - bulk_salary_start

        if not bulk_salary_result['success']:
            return jsonify({
                'success': False,
                'message': f'Bulk salary calculation failed: {bulk_salary_result.get("message", "Unknown error")}'
            }), 500

        salary_data_list = bulk_salary_result['data']
        salary_data_dict = {item['Employee ID']: item for item in salary_data_list}

        # ============================================
        # STEP 3: BULK fetch attendance data for ALL employees (1 query)
        # ============================================
        bulk_attendance_start = time.time()
        from services.attendance_service import AttendanceService
        bulk_attendance_result = AttendanceService.get_bulk_monthly_attendance_summary(employee_ids, year, month)
        bulk_attendance_time = time.time() - bulk_attendance_start

        if not bulk_attendance_result['success']:
            return jsonify({
                'success': False,
                'message': f'Bulk attendance calculation failed: {bulk_attendance_result.get("message", "Unknown error")}'
            }), 500

        attendance_dict = bulk_attendance_result['data']

        # ============================================
        # STEP 4: BULK fetch wage master data for ALL employees (1 query)
        # ============================================
        wage_master_start = time.time()
        wage_master_query = WageMaster.query.filter(
            WageMaster.salary_code.in_([emp.salary_code for emp in employees if emp.salary_code])
        ).all()

        wage_master_dict = {wm.salary_code: wm for wm in wage_master_query}
        wage_master_time = time.time() - wage_master_start

        # ============================================
        # STEP 5: Build Form D rows with ESIC calculations (IN-MEMORY)
        # ============================================
        form_d_data = []

        for idx, employee in enumerate(employees, 1):
            salary_data = salary_data_dict.get(employee.employee_id, {})
            attendance_data = attendance_dict.get(employee.employee_id, {})

            wage_master = wage_master_dict.get(employee.salary_code)

            # Get monthly basic wage (not total earnings)
            total_monthly_wages = salary_data.get('Basic', 0)

            # Check if employee is covered under ESIC (wage ceiling)
            if total_monthly_wages > ESIC_WAGE_CEILING:
                continue  # Skip employees above wage ceiling

            # Get number of days worked
            no_of_days = attendance_data.get('present_days', 0)

            # Calculate IP contribution (0.75% of total monthly wages, rounded up)
            ip_contribution = max(1, round(total_monthly_wages * 0.0075, 0))  # ROUNDUP equivalent

            # Fetch insurance number from employee's stored ESIC number
            try:
                if employee.esic_number and employee.esic_number.strip():
                    insurance_no = employee.esic_number.strip()
                else:
                    insurance_no = "Null"
            except AttributeError:
                # Handle case where esic_number field doesn't exist or is None
                insurance_no = "Null"
            except Exception as e:
                # Handle any other potential errors gracefully
                print(f"[WARNING] Error fetching ESIC number for employee {employee.employee_id}: {str(e)}")
                insurance_no = "Null"

            form_d_row = {
                "slNo": idx,
                "insuranceNo": insurance_no,
                "nameOfInsuredPerson": f"{employee.first_name} {employee.last_name}",
                "noOfDays": no_of_days,
                "totalMonthlyWages": total_monthly_wages,
                "ipContribution": ip_contribution,
                "siteName": wage_master.site_name if wage_master else "N/A"
            }

            form_d_data.append(form_d_row)

        # Calculate totals
        totals = {
            "totalEmployees": len(form_d_data),
            "totalDays": sum(row["noOfDays"] for row in form_d_data),
            "totalMonthlyWages": sum(row["totalMonthlyWages"] for row in form_d_data),
            "totalIpContribution": sum(row["ipContribution"] for row in form_d_data)
        }

        # Calculate performance metrics
        total_time = time.time() - start_time

        response = jsonify({
            "success": True,
            "data": form_d_data,
            "totals": totals,
            "filters": {
                "year": year,
                "month": month,
                "site": site,
                "monthName": calendar.month_name[month]
            },
            "performance": {
                "total_time_seconds": round(total_time, 3),
                "employees_processed": len(employees),
                "bulk_salary_time": round(bulk_salary_time, 3),
                "bulk_attendance_time": round(bulk_attendance_time, 3),
                "wage_master_time": round(wage_master_time, 3),
                "total_queries": 4,
                "old_queries_estimate": len(employees) * 8
            }
        })

        # Add performance headers
        response.headers['X-Performance-Total'] = f'{total_time:.3f}s'
        response.headers['X-Performance-Employees'] = str(len(employees))
        response.headers['X-Performance-Queries'] = '4'

        return response, 200

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error generating Form D data: {str(e)}"
        }), 500


@forms_bp.route("/form-d/download", methods=["GET", "OPTIONS"])
def download_form_d_excel():
    """
    Download Form D data as Excel file
    """
    # Handle preflight OPTIONS request
    if request.method == 'OPTIONS':
        return '', 200

    try:
        # Get query parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        site = request.args.get('site', 'All')

        if not year or not month:
            return jsonify({
                "success": False,
                "message": "Year and month are required"
            }), 400

        # Get Form D data using the same logic as the data endpoint
        form_d_response = get_form_d_data()

        if form_d_response[1] != 200:
            return form_d_response

        form_d_result = form_d_response[0].get_json()
        form_d_data = form_d_result['data']
        totals = form_d_result['totals']

        # Create Excel file
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Prepare data for Excel
            excel_data = []

            for row in form_d_data:
                excel_row = {
                    'Sl. No.': row['slNo'],
                    'Insurance No.': row['insuranceNo'],
                    'Name of the Insured Person': row['nameOfInsuredPerson'],
                    'No. Of Days': row['noOfDays'],
                    'Total Monthly Wages': row['totalMonthlyWages'],
                    'IP Contribution': row['ipContribution']
                }
                excel_data.append(excel_row)

            # Add totals row
            totals_row = {
                'Sl. No.': '',
                'Insurance No.': '',
                'Name of the Insured Person': 'TOTAL',
                'No. Of Days': totals['totalDays'],
                'Total Monthly Wages': totals['totalMonthlyWages'],
                'IP Contribution': totals['totalIpContribution']
            }
            excel_data.append(totals_row)

            # Create DataFrame and write to Excel
            df = pd.DataFrame(excel_data)
            df.to_excel(writer, sheet_name='Form D - ESIC', index=False)

            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Form D - ESIC']

            # Add header information
            worksheet.insert_rows(1, 5)
            worksheet['A1'] = "Form D"
            worksheet['A2'] = "Employees' State Insurance Corporation (ESIC)"
            worksheet['A3'] = f"Wage Ceiling: ₹21,000 | Date: {datetime.now().strftime('%d/%m/%Y')}"
            worksheet['A4'] = "SSPL"
            worksheet['A5'] = f"Month: {calendar.month_name[month]} {year} | Site: {site}"

            # Style the headers
            from openpyxl.styles import Font, Alignment

            for row in range(1, 6):
                cell = worksheet[f'A{row}']
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center')

            # Merge cells for headers
            worksheet.merge_cells('A1:F1')
            worksheet.merge_cells('A2:F2')
            worksheet.merge_cells('A3:F3')
            worksheet.merge_cells('A4:F4')
            worksheet.merge_cells('A5:F5')

        output.seek(0)

        # Generate filename
        month_name = calendar.month_name[month]
        filename = f"FormD_ESIC_{site}_{month_name}_{year}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error generating Excel file: {str(e)}"
        }), 500


@forms_bp.route("/form-c", methods=["GET", "OPTIONS"])
def get_form_c_data():
    """
    Get Form C (EPF) data with wage ceilings and contribution calculations
    """
    # Handle preflight OPTIONS request
    if request.method == 'OPTIONS':
        return '', 200

    import time
    start_time = time.time()

    try:
        # Get query parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        site = request.args.get('site')

        if not year or not month:
            return jsonify({
                "success": False,
                "message": "Year and month are required"
            }), 400

        # EPF wage ceiling (₹15,000)
        EPF_WAGE_CEILING = 15000

        # ============================================
        # STEP 1: Get employees for the specified site (1 query)
        # ============================================
        query = Employee.query
        if site:
            # URL decode the site parameter
            from urllib.parse import unquote
            decoded_site = unquote(site)
            print(f"[DEBUG] Filtering by site: '{site}' -> '{decoded_site}'")

            # Join with WageMaster to filter by site
            query = query.join(WageMaster, Employee.salary_code == WageMaster.salary_code)\
                         .filter(WageMaster.site_name == decoded_site)

        # Apply sorting and execute query
        employees = query.order_by(Employee.employee_id.asc()).all()

        if not employees:
            return jsonify({
                "success": False,
                "message": "No employees found for the specified criteria"
            }), 404

        employee_ids = [emp.employee_id for emp in employees]

        # ============================================
        # STEP 2: BULK calculate salaries for filtered employees (3 queries)
        # ============================================
        bulk_salary_start = time.time()
        bulk_salary_result = SalaryService.generate_monthly_salary_data_for_employees(employee_ids, year, month)
        bulk_salary_time = time.time() - bulk_salary_start

        if not bulk_salary_result['success']:
            return jsonify({
                'success': False,
                'message': f'Bulk salary calculation failed: {bulk_salary_result.get("message", "Unknown error")}'
            }), 500

        salary_data_list = bulk_salary_result['data']
        salary_data_dict = {item['Employee ID']: item for item in salary_data_list}

        # ============================================
        # STEP 3: BULK fetch attendance data for ALL employees (1 query)
        # ============================================
        bulk_attendance_start = time.time()
        from services.attendance_service import AttendanceService
        bulk_attendance_result = AttendanceService.get_bulk_monthly_attendance_summary(employee_ids, year, month)
        bulk_attendance_time = time.time() - bulk_attendance_start

        if not bulk_attendance_result['success']:
            return jsonify({
                'success': False,
                'message': f'Bulk attendance calculation failed: {bulk_attendance_result.get("message", "Unknown error")}'
            }), 500

        attendance_dict = bulk_attendance_result['data']

        # ============================================
        # STEP 4: BULK fetch wage master data for ALL employees (1 query)
        # ============================================
        wage_master_start = time.time()
        wage_master_query = WageMaster.query.filter(
            WageMaster.salary_code.in_([emp.salary_code for emp in employees if emp.salary_code])
        ).all()

        wage_master_dict = {wm.salary_code: wm for wm in wage_master_query}
        wage_master_time = time.time() - wage_master_start

        # ============================================
        # STEP 5: Build Form C rows with EPF calculations (IN-MEMORY)
        # ============================================
        form_c_data = []

        for idx, employee in enumerate(employees, 1):
            salary_data = salary_data_dict.get(employee.employee_id, {})
            attendance_data = attendance_dict.get(employee.employee_id, {})

            wage_master = wage_master_dict.get(employee.salary_code)

            # Get basic salary (gross wages)
            gross_wages = salary_data.get('Basic', 0)

            # Apply EPF wage ceiling for EPF, EPS, EDLI wages
            epf_wages = min(gross_wages, EPF_WAGE_CEILING)
            eps_wages = min(gross_wages, EPF_WAGE_CEILING)
            edli_wages = min(gross_wages, EPF_WAGE_CEILING)

            # Calculate contributions
            epf_contribution = round(epf_wages * 0.12, 0)  # 12% EPF contribution
            eps_contribution = round(eps_wages * 0.0833, 0)  # 8.33% EPS contribution
            epf_eps_diff = epf_contribution - eps_contribution

            # NCP Days (Non-Contributing Period) - for now, assume 0
            ncp_days = 0

            # Refund of Advance - for now, assume 0
            refund_of_advance = 0

            form_c_row = {
                "slNo": idx,
                "memberName": f"{employee.first_name} {employee.last_name}",
                "grossWages": gross_wages,
                "epfWages": epf_wages,
                "epsWages": eps_wages,
                "edliWages": edli_wages,
                "epfContribution": epf_contribution,
                "epsContribution": eps_contribution,
                "epfEpsDiff": epf_eps_diff,
                "ncpDays": ncp_days,
                "refundOfAdvance": refund_of_advance,
                "siteName": wage_master.site_name if wage_master else "N/A"
            }

            form_c_data.append(form_c_row)

        # Calculate totals
        totals = {
            "totalEmployees": len(form_c_data),
            "totalGrossWages": sum(row["grossWages"] for row in form_c_data),
            "totalEpfWages": sum(row["epfWages"] for row in form_c_data),
            "totalEpsWages": sum(row["epsWages"] for row in form_c_data),
            "totalEdliWages": sum(row["edliWages"] for row in form_c_data),
            "totalEpfContribution": sum(row["epfContribution"] for row in form_c_data),
            "totalEpsContribution": sum(row["epsContribution"] for row in form_c_data),
            "totalEpfEpsDiff": sum(row["epfEpsDiff"] for row in form_c_data),
            "totalNcpDays": sum(row["ncpDays"] for row in form_c_data),
            "totalRefundOfAdvance": sum(row["refundOfAdvance"] for row in form_c_data)
        }

        # Calculate performance metrics
        total_time = time.time() - start_time

        response = jsonify({
            "success": True,
            "data": form_c_data,
            "totals": totals,
            "filters": {
                "year": year,
                "month": month,
                "site": site,
                "monthName": calendar.month_name[month]
            },
            "performance": {
                "total_time_seconds": round(total_time, 3),
                "employees_processed": len(employees),
                "bulk_salary_time": round(bulk_salary_time, 3),
                "bulk_attendance_time": round(bulk_attendance_time, 3),
                "wage_master_time": round(wage_master_time, 3),
                "total_queries": 4,
                "old_queries_estimate": len(employees) * 8
            }
        })

        # Add performance headers
        response.headers['X-Performance-Total'] = f'{total_time:.3f}s'
        response.headers['X-Performance-Employees'] = str(len(employees))
        response.headers['X-Performance-Queries'] = '4'

        return response, 200

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error generating Form C data: {str(e)}"
        }), 500


@forms_bp.route("/form-c/download", methods=["GET", "OPTIONS"])
def download_form_c_excel():
    """
    Download Form C data as Excel file
    """
    # Handle preflight OPTIONS request
    if request.method == 'OPTIONS':
        return '', 200

    try:
        # Get query parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        site = request.args.get('site', 'All')

        if not year or not month:
            return jsonify({
                "success": False,
                "message": "Year and month are required"
            }), 400

        # Get Form C data using the same logic as the data endpoint
        form_c_response = get_form_c_data()

        if form_c_response[1] != 200:
            return form_c_response

        form_c_result = form_c_response[0].get_json()
        form_c_data = form_c_result['data']
        totals = form_c_result['totals']

        # Create Excel file
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Prepare data for Excel
            excel_data = []

            for row in form_c_data:
                excel_row = {
                    'Sl.No': row['slNo'],
                    'Member Name': row['memberName'],
                    'Gross Wages': row['grossWages'],
                    'EPF Wages': row['epfWages'],
                    'EPS Wages': row['epsWages'],
                    'EDLI Wages': row['edliWages'],
                    'EPF Contribution': row['epfContribution'],
                    'EPS Contribution': row['epsContribution'],
                    'EPF EPS Diff': row['epfEpsDiff'],
                    'NCP Days': row['ncpDays'],
                    'Refund of Advance': row['refundOfAdvance']
                }
                excel_data.append(excel_row)

            # Add totals row
            totals_row = {
                'Sl.No': '',
                'Member Name': 'TOTAL',
                'Gross Wages': totals['totalGrossWages'],
                'EPF Wages': totals['totalEpfWages'],
                'EPS Wages': totals['totalEpsWages'],
                'EDLI Wages': totals['totalEdliWages'],
                'EPF Contribution': totals['totalEpfContribution'],
                'EPS Contribution': totals['totalEpsContribution'],
                'EPF EPS Diff': totals['totalEpfEpsDiff'],
                'NCP Days': totals['totalNcpDays'],
                'Refund of Advance': totals['totalRefundOfAdvance']
            }
            excel_data.append(totals_row)

            # Create DataFrame and write to Excel
            df = pd.DataFrame(excel_data)
            df.to_excel(writer, sheet_name='Form C - EPF', index=False)

            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Form C - EPF']

            # Add header information
            worksheet.insert_rows(1, 5)
            worksheet['A1'] = "Form C"
            worksheet['A2'] = "Employees' Provident Fund (EPF)"
            worksheet['A3'] = f"Wage Ceiling: ₹15,000 | Date: {datetime.now().strftime('%d/%m/%Y')}"
            worksheet['A4'] = "SSPL"
            worksheet['A5'] = f"Month: {calendar.month_name[month]} {year} | Site: {site}"

            # Style the headers
            from openpyxl.styles import Font, Alignment

            for row in range(1, 6):
                cell = worksheet[f'A{row}']
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center')

            # Merge cells for headers
            worksheet.merge_cells('A1:K1')
            worksheet.merge_cells('A2:K2')
            worksheet.merge_cells('A3:K3')
            worksheet.merge_cells('A4:K4')
            worksheet.merge_cells('A5:K5')

        output.seek(0)

        # Generate filename
        month_name = calendar.month_name[month]
        filename = f"FormC_EPF_{site}_{month_name}_{year}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error generating Excel file: {str(e)}"
        }), 500
